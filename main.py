
#  Python imports
import xlsxwriter 
import xlsxwriter 
import pandas as pd
import openpyxl as op
from numpy import nan
import openpyxl as op
from time import sleep
from bs4 import BeautifulSoup

# selenium imports
from selenium import webdriver
from random_user_agent.user_agent import UserAgent
from selenium.webdriver.chrome.options import Options
from random_user_agent.params import SoftwareName, OperatingSystem


class octopart:
    filename = 'output.xlsx'
    browser = None

    def __init__(self):
        self.create_excel_file()
        self.digikey_login()

    def scrape(self, partnumber):
        # print('\n SCRAPING  . . . \n')
        # self.browser = webdriver.Chrome('chromedriver.exe') 
        url = 'https://octopart.com/search?q=' + partnumber + '&currency=USD&specs=0'
        self.browser.get(url)
        bs = self.browser.find_elements_by_tag_name("button")  
        for b in bs: 
            if b.text == 'Show All': b.click()
        soup = BeautifulSoup(self.browser.page_source, features='lxml')
        table = soup.find('table')
        tbody = table.find('tbody')
        trs = tbody.find_all('tr')
        print('\n\n')
        for tr in trs:
            if str(tr.get_text()).__contains__('Digi-Key') :
                stock = tr.find('td', attrs={'class': 'jsx-3779711368'})
                moq = tr.find('td', attrs={'class': 'jsx-3318793868'})
                price = tr.find('td', attrs={'class': 'jsx-1795508439'})
                sku = tr.find('td', attrs={'class': 'jsx-2389699081 sku'})
                moq = moq.get_text()
                sku = sku.find('a').get_text()
                stock = str(stock.get_text()).replace(',','')
                try: price = float(price.get_text())
                except: price = 'None'
                desc = 'None'
                try:
                    link = tr.find('a', attrs={'class': 'jsx-2100737765 click-url'})['href']
                    self.browser.get(link)
                    desc = self.digikey(self.browser.page_source)
                except: 
                    sleep(15)
                    try:
                        link = tr.find('a', attrs={'class': 'jsx-2100737765 click-url'})['href']
                        self.browser.get(link)
                        desc = self.digikey(self.browser.page_source)
                    except: pass
                
                print('Digi-Key:')
                print('\tpartID = ', partnumber)
                print('\tS Code = ', sku)
                print('\tMOQ    = ', moq)
                print('\tStock  = ', int(stock))
                print('\tPrice  = ', price)
                print('\tDesc   = ', desc)
                self.write_to_excel(partnumber, desc, 'Digi-Key', link, sku, stock, moq, price)
                
            # if str(tr.get_text()).__contains__('Mouser') :
                # td = tr.find('td', attrs={'class': 'jsx-3779711368'})
                # value = str(td.get_text()).replace(',','')
                # td2 = tr.find('td', attrs={'class': 'jsx-1795508439'})
                # try: price = float(td2.get_text())
                # except: price = 'None'       
                # print('Mouser:')
                # print('\tpartID = ', partnumber)
                # print('\tStock = ', int(value))
                # print('\tPrice = ', price)
                # self.write_to_excel(partnumber, 'Mouser', value, price)

            # if str(tr.get_text()).__contains__('Farnell') :
                # td = tr.find('td', attrs={'class': 'jsx-3779711368'})
                # value = str(td.get_text()).replace(',','')
                # td2 = tr.find('td', attrs={'class': 'jsx-1795508439'})
                # try: price = float(td2.get_text())
                # except: price = 'None'
                # print('Farnell:')
                # print('\tpartID = ', partnumber)
                # print('\tStock = ', int(value))
                # print('\tPrice = ', price)
                # self.write_to_excel(partnumber, 'Farnell', value, price)
    
    def close_browser(self):
        self.browser.quit()

    def digikey_login(self):
        login_url = 'https://auth.digikey.com/as/authorization.oauth2?response_type=code&client_id=pa_wam&redirect_uri=https%3A%2F%2Fwww.digikey.com%2Fpa%2Foidc%2Fcb&state=eyJ6aXAiOiJERUYiLCJhbGciOiJkaXIiLCJlbmMiOiJBMTI4Q0JDLUhTMjU2Iiwia2lkIjoiMnMiLCJzdWZmaXgiOiJqMlF6T0guMTYxNjE1NjE1MCJ9..2GHWxHGWNbzHWGmhnZy-bg.mOYu2qAf3Dpw5XtEa9k-T9dAkymsZddvUugkoGWVZC_AtbxvCJcGU1p6T6nW9tDQPZ5EhrOs5T5EPtDLhcRepDN0G9_bVYIvlsQYzBsk9gc.GA5-rxU38vBdh4ZUJwVIww&nonce=OzIsYKw5g25-48HBJ9SdSDok-fYrnuGMm1-_hR03u7w&scope=openid%20address%20email%20phone%20profile&vnd_pi_requested_resource=https%3A%2F%2Fwww.digikey.com%2FMyDigiKey&vnd_pi_application_name=DigikeyProd-Mydigikey'
        try:
            software_names = [SoftwareName.CHROME.value]
            operating_systems = [OperatingSystem.WINDOWS.value, OperatingSystem.LINUX.value]
            user_agent_rotator = UserAgent(software_names=software_names, operating_systems=operating_systems, limit=100)
            user_agent = user_agent_rotator.get_random_user_agent()

            chrome_options = Options()           
            chrome_options.add_argument(f'user-agent={user_agent}')
            # chrome_options.add_argument('--disable-extensions')
            # chrome_options.add_argument('--profile-directory=Default')
            # chrome_options.add_argument("--disable-plugins-discovery")
            # chrome_options.add_argument("--start-maximized")
            # chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            # chrome_options.add_experimental_option('useAutomationExtension', False)
            # chrome_options.add_argument('--disable-blink-features=AutomationControlled')

            driver = webdriver.Chrome(options=chrome_options)
            driver.delete_all_cookies()
            driver.set_window_position(0,0)
            driver.maximize_window()
            self.browser = driver
            self.browser.get(login_url)
            elementID = self.browser.find_element_by_id('username')
            elementID.send_keys("adnanshahz2015@gmail.com")
            elementID = self.browser.find_element_by_id('password')
            elementID.send_keys("zeshan2015")
            elementID.submit()

            self.browser.get('https://digikey.com/')
            self.browser.find_element_by_id('my_digikey_logged_in').click()
            sleep(7)
        except: 
            print("\nCan't Login to DigiKey - - -\n")

    def farnell_login(self):   
        login_url = 'https://uk.farnell.com/webapp/wcs/stores/servlet/LogonForm?myAcctMain=1&catalogId=15001&storeId=10151&langId=44'
        self.browser = webdriver.Chrome('chromedriver.exe') 
        self.browser.get(login_url)
        username = self.browser.find_element_by_id("logonId")
        password = self.browser.find_element_by_id("logonPassword")
        username.send_keys("adnanxshah")
        password.send_keys("Zeshan2015-")
        login = self.browser.find_element_by_link_text('Log In')
        login.click()
        print(login)
        sleep(5)

    def digikey(self, source=None):
        soup = BeautifulSoup(source, features='lxml')
        tbody = soup.find('tbody', attrs={'class' : 'MuiTableBody-root'})
        tr = tbody.find_all('tr')[4]
        print('tr = ', tr)
        td = tr.find_all('td')[1]
        print('td == ', td)
        desc = td.find('div').get_text()
        return desc

    def create_excel_file(self):
        # creating the file for the first time 
        # if not os.path.exists(self.filename):
        workbook = xlsxwriter.Workbook(self.filename)
        worksheet = workbook.add_worksheet("data")
        workbook.close()
        wb = op.load_workbook(self.filename, False)
        ws = wb['data']
        ws.append(['Manufacturing Part Number', 'Description', 'Distributor', 'Web URL', 'Supplier Code', 'Stock', 'MOQ', 'Price'])
        wb.save(self.filename)
        wb.close()

    def write_to_excel(self, partnumber, desc, distributor, web_url, supplier_code, stock, moq, price):
        wb = op.load_workbook(self.filename, False)
        ws = wb['data']
        ws.append([partnumber, desc, distributor, web_url, supplier_code, stock, moq, price])
        wb.save(self.filename)
        wb.close()


def selected_part_numbers():
    part_list = []
    file_data =  pd.read_excel('input.xlsx', 'data')
    datas = file_data['manufacture part number']
    for i in datas.index:
        part_list.append(datas[i])
    return part_list

if __name__ == '__main__':
    obj = octopart()
    part_list = selected_part_numbers() 
    print(part_list)   
    for part in part_list:
        obj.scrape(part)
        sleep(7)
    obj.close_browser()

